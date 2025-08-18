import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment


def to_float_safe(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def generate_calibration_report(results: dict, output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "보정값"

    headers = ["SN", "pm2.5", "pm10", "temp", "humi", "co2"]
    ws.append(headers)

    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 헤더 셀 스타일 적용
    for cell in ws[1]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    for file_path, result in results.items():
        sn = os.path.splitext(os.path.basename(file_path))[0]

        pm25_formula = ",".join([f"*{v}" for _, v in result.get("pm25_correction", [])])
        pm10_formula = ",".join([f"*{v}" for _, v in result.get("pm10_correction", [])])

        temp_corr_raw = round(to_float_safe(result.get("temp_correction")), 1)
        humi_corr_raw = round(to_float_safe(result.get("humi_correction")), 1)

        temp_corr = f"{temp_corr_raw:+.1f}"
        humi_corr = f"{humi_corr_raw:+.1f}"

        co2_corr = result.get("co2_correction_str", "")

        row_header = [sn, pm25_formula, pm10_formula, temp_corr, humi_corr, co2_corr]
        ws.append(row_header)

        row_pre = [
            "보정 전",
            f"{to_float_safe(result.get('pm25_accuracy_pre')):.0f}%",
            f"{to_float_safe(result.get('pm10_accuracy_pre')):.0f}%",
            f"{to_float_safe(result.get('temp_accuracy')):.0f}%",
            f"{to_float_safe(result.get('humi_accuracy')):.0f}%",
            f"{to_float_safe(result.get('pre_correction_accuracy')):.0f}%"
        ]
        ws.append(row_pre)

        row_post = [
            "보정 후",
            f"{to_float_safe(result.get('pm25_accuracy_post')):.0f}%",
            f"{to_float_safe(result.get('pm10_accuracy_post')):.0f}%",
            f"{to_float_safe(result.get('temp_corrected_accuracy')):.0f}%",
            f"{to_float_safe(result.get('humi_corrected_accuracy')):.0f}%",
            f"{to_float_safe(result.get('post_correction_accuracy')):.0f}%"
        ]
        ws.append(row_post)

        ws.append([])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_align

    wb.save(output_file)
